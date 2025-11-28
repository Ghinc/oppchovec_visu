"""
Régénère la CAH 3D complète avec:
- Couleurs cohérentes entre cartes et Excel
- Légende simplifiée (juste n communes)
- Tous les outputs: cartes, Excel, JSON
"""

import json
import pandas as pd
import geopandas as gpd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
from scipy.cluster.hierarchy import linkage, dendrogram, fcluster
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import silhouette_score, davies_bouldin_score
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Palette de couleurs fixe pour les clusters (8 couleurs distinctes)
CLUSTER_COLORS_HEX = [
    "#E41A1C",  # Rouge
    "#377EB8",  # Bleu
    "#4DAF4A",  # Vert
    "#984EA3",  # Violet
    "#FF7F00",  # Orange
    "#FFFF33",  # Jaune
    "#A65628",  # Marron
    "#F781BF"   # Rose
]

# Pour Excel (sans #)
CLUSTER_COLORS_EXCEL = [c[1:] for c in CLUSTER_COLORS_HEX]

def charger_donnees(json_path, geojson_path):
    """Charge les données"""
    print("[INFO] Chargement des donnees...")

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    gdf = gpd.read_file(geojson_path)

    print(f"  [OK] {len(data)} communes chargees depuis JSON")
    print(f"  [OK] {len(gdf)} communes chargees depuis GeoJSON")

    return data, gdf

def preparer_donnees(data, gdf):
    """Prépare les données"""
    print("\n[INFO] Preparation des donnees pour CAH 3D...")

    df_scores = pd.DataFrame.from_dict(data, orient='index')
    df_scores.index.name = 'commune'
    df_scores.reset_index(inplace=True)

    gdf['nom_clean'] = gdf['nom'].str.strip()
    gdf_merged = gdf.merge(df_scores, left_on='nom_clean', right_on='commune', how='inner')

    print(f"  [OK] {len(gdf_merged)} communes fusionnees avec succes")

    X = gdf_merged[['Score_Opp_0_10', 'Score_Cho_0_10', 'Score_Vec_0_10']].values

    print(f"  [INFO] Matrice de donnees: {X.shape}")
    print(f"  [INFO] Score_Opp: Min={X[:, 0].min():.2f}, Max={X[:, 0].max():.2f}, Moyenne={X[:, 0].mean():.2f}")
    print(f"  [INFO] Score_Cho: Min={X[:, 1].min():.2f}, Max={X[:, 1].max():.2f}, Moyenne={X[:, 1].mean():.2f}")
    print(f"  [INFO] Score_Vec: Min={X[:, 2].min():.2f}, Max={X[:, 2].max():.2f}, Moyenne={X[:, 2].mean():.2f}")

    return gdf_merged, X

def calculer_cah(X):
    """Calcule la CAH"""
    print("\n[INFO] Calcul de la CAH 3D (methode Ward, distance euclidienne)...")

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    Z = linkage(X_scaled, method='ward')

    print(f"  [OK] Matrice de liaison calculee")

    return X_scaled, Z

def generer_dendrogramme(Z, output_path):
    """Génère le dendrogramme"""
    print(f"\n[INFO] Generation du dendrogramme: {output_path}")

    plt.figure(figsize=(20, 10))
    dendrogram(Z, labels=None, leaf_font_size=8, no_labels=True)
    plt.title('Dendrogramme - Classification Hierarchique Ascendante (3D: Opp, Cho, Vec)\nMethode: Ward | Distance: Euclidienne',
              fontsize=16, weight='bold', pad=20)
    plt.xlabel('Communes', fontsize=14)
    plt.ylabel('Distance', fontsize=14)
    plt.grid(True, alpha=0.3, axis='y')
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  [OK] Dendrogramme sauvegarde")

def calculer_tous_clusters(Z, n_max=8):
    """Calcule les clusters"""
    print(f"\n[INFO] Calcul des clusters pour n=range(2, {n_max + 1})...")

    resultats = {}
    for n in range(2, n_max + 1):
        clusters = fcluster(Z, t=n, criterion='maxclust')
        resultats[n] = clusters
        print(f"  [OK] {n} clusters: {len(np.unique(clusters))} groupes trouves")

    return resultats

def calculer_metriques_qualite(X_scaled, clusters_dict):
    """Calcule les métriques de qualité"""
    print("\n[INFO] Calcul des metriques de qualite...")

    metriques = []
    for n, clusters in clusters_dict.items():
        silhouette = silhouette_score(X_scaled, clusters)
        davies_bouldin = davies_bouldin_score(X_scaled, clusters)

        inertie_intra = 0.0
        for i in range(1, n + 1):
            cluster_points = X_scaled[clusters == i]
            if len(cluster_points) > 0:
                centroid = cluster_points.mean(axis=0)
                inertie_intra += np.sum((cluster_points - centroid) ** 2)

        inertie_totale = np.sum((X_scaled - X_scaled.mean(axis=0)) ** 2)
        pct_variance_expliquee = ((inertie_totale - inertie_intra) / inertie_totale) * 100

        metriques.append({
            'n_clusters': n,
            'silhouette': silhouette,
            'davies_bouldin': davies_bouldin,
            'inertie_intra': inertie_intra,
            'variance_expliquee': pct_variance_expliquee
        })

        print(f"  [OK] {n} clusters: Silhouette={silhouette:.4f}, Davies-Bouldin={davies_bouldin:.4f}, "
              f"Inertie={inertie_intra:.2f}, Variance={pct_variance_expliquee:.1f}%")

    return pd.DataFrame(metriques)

def generer_cartes_clusters(gdf, clusters_dict, output_dir):
    """Génère les cartes avec couleurs cohérentes"""
    print("\n[INFO] Generation des cartes de clusters...")

    for n, clusters in clusters_dict.items():
        gdf_temp = gdf.copy()
        gdf_temp['cluster'] = clusters

        # Créer une colormap personnalisée avec nos couleurs
        cmap = ListedColormap(CLUSTER_COLORS_HEX[:n])

        fig, ax = plt.subplots(1, 1, figsize=(12, 14))

        # Plot avec couleurs personnalisées
        gdf_temp.plot(column='cluster', ax=ax, legend=True, cmap=cmap,
                     edgecolor='black', linewidth=0.5,
                     categorical=True,
                     legend_kwds={'loc': 'lower left', 'fontsize': 10})

        # Modifier la légende pour afficher seulement le nombre de communes
        legend = ax.get_legend()
        if legend:
            # Calculer le nombre de communes par cluster
            for i, text in enumerate(legend.get_texts()):
                cluster_id = i + 1
                n_communes = len(gdf_temp[gdf_temp['cluster'] == cluster_id])
                text.set_text(f'n={n_communes}')

        ax.set_title(f'Classification Hierarchique Ascendante - {n} clusters\n'
                    f'Variables: Opp, Cho, Vec | Methode: Ward',
                    fontsize=14, weight='bold', pad=15)
        ax.axis('off')

        output_path = f"{output_dir}/cah_3d_carte_{n}_clusters.png"
        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()

        print(f"  [OK] Carte sauvegardee: cah_3d_carte_{n}_clusters.png")

def analyser_clusters(gdf, clusters_dict):
    """Analyse statistique des clusters"""
    print("\n[INFO] Analyse statistique des clusters...")

    analyses = {}
    for n, clusters in clusters_dict.items():
        gdf_temp = gdf.copy()
        gdf_temp['cluster'] = clusters

        stats = []
        for i in range(1, n + 1):
            cluster_data = gdf_temp[gdf_temp['cluster'] == i]
            stats.append({
                'cluster': i,
                'n_communes': len(cluster_data),
                'opp_moyenne': cluster_data['Score_Opp_0_10'].mean(),
                'cho_moyenne': cluster_data['Score_Cho_0_10'].mean(),
                'vec_moyenne': cluster_data['Score_Vec_0_10'].mean(),
                'oppchovec_moyenne': cluster_data['OppChoVec_0_10'].mean(),
                'opp_std': cluster_data['Score_Opp_0_10'].std(),
                'cho_std': cluster_data['Score_Cho_0_10'].std(),
                'vec_std': cluster_data['Score_Vec_0_10'].std()
            })

        analyses[n] = pd.DataFrame(stats)

    return analyses

def exporter_vers_excel(metriques_df, analyses_dict, output_path):
    """Exporte vers Excel avec couleurs cohérentes"""
    print(f"\n[INFO] Export vers Excel: {output_path}")

    wb = Workbook()

    # Feuille 1 : Métriques de qualité
    ws1 = wb.active
    ws1.title = "Metriques de qualite"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    ws1['A1'] = "Metriques de qualite - CAH 3D (Opp, Cho, Vec)"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:F1')

    headers = ['Nombre de clusters', 'Silhouette Score', 'Davies-Bouldin Index',
               'Inertie intra-classe', 'Variance expliquee (%)']
    for col, header in enumerate(headers, start=1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for idx, row in metriques_df.iterrows():
        ws1.cell(row=idx+4, column=1, value=row['n_clusters']).border = border
        ws1.cell(row=idx+4, column=2, value=round(row['silhouette'], 4)).border = border
        ws1.cell(row=idx+4, column=3, value=round(row['davies_bouldin'], 4)).border = border
        ws1.cell(row=idx+4, column=4, value=round(row['inertie_intra'], 2)).border = border
        ws1.cell(row=idx+4, column=5, value=round(row['variance_expliquee'], 2)).border = border

    best_silhouette = metriques_df.loc[metriques_df['silhouette'].idxmax()]
    ws1[f'A{len(metriques_df)+6}'] = f"Meilleur Silhouette Score: {best_silhouette['n_clusters']} clusters"
    ws1[f'A{len(metriques_df)+6}'].font = Font(bold=True, color="00B050")

    # Feuilles suivantes : Statistiques par nombre de clusters (avec couleurs cohérentes)
    for n, stats_df in analyses_dict.items():
        ws = wb.create_sheet(title=f"{n} clusters")

        ws['A1'] = f"Analyse des {n} clusters - Variables: Opp, Cho, Vec"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:J1')

        headers_stats = ['Cluster', 'N communes', 'Opp moyen', 'Cho moyen', 'Vec moyen',
                        'OppChoVec moyen', 'Opp std', 'Cho std', 'Vec std']
        for col, header in enumerate(headers_stats, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        for idx, row in stats_df.iterrows():
            cluster_id = int(row['cluster'])

            # Appliquer la couleur du cluster à toute la ligne
            for col in range(1, 10):
                cell = ws.cell(row=idx+4, column=col)
                cell.border = border
                cell.fill = PatternFill(
                    start_color=CLUSTER_COLORS_EXCEL[cluster_id - 1],
                    end_color=CLUSTER_COLORS_EXCEL[cluster_id - 1],
                    fill_type="solid"
                )
                cell.font = Font(bold=True, color="000000")  # Texte noir pour meilleure lisibilité

            # Remplir les valeurs
            ws.cell(row=idx+4, column=1, value=row['cluster'])
            ws.cell(row=idx+4, column=2, value=row['n_communes'])
            ws.cell(row=idx+4, column=3, value=round(row['opp_moyenne'], 2))
            ws.cell(row=idx+4, column=4, value=round(row['cho_moyenne'], 2))
            ws.cell(row=idx+4, column=5, value=round(row['vec_moyenne'], 2))
            ws.cell(row=idx+4, column=6, value=round(row['oppchovec_moyenne'], 2))
            ws.cell(row=idx+4, column=7, value=round(row['opp_std'], 2))
            ws.cell(row=idx+4, column=8, value=round(row['cho_std'], 2))
            ws.cell(row=idx+4, column=9, value=round(row['vec_std'], 2))

    wb.save(output_path)
    print(f"  [OK] Fichier Excel cree")

def exporter_clusters_communes(gdf, clusters_dict, output_path):
    """Exporte Excel avec appartenance des communes"""
    print(f"\n[INFO] Export clusters par commune: {output_path}")

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for n in range(2, 9):
        ws = wb.create_sheet(title=f"{n} clusters")

        # Titre
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Classification Hierarchique Ascendante 3D - {n} clusters"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:F2')
        ws['A2'] = f"Variables: Opp, Cho, Vec | Methode: Ward | Distance: Euclidienne"
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:F3')
        ws['A3'] = f"Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M:%S')}"
        ws['A3'].font = Font(italic=True, size=9)
        ws['A3'].alignment = Alignment(horizontal='center')

        # En-têtes
        headers = ['Commune', 'Cluster', 'Opp', 'Cho', 'Vec', 'OppChoVec']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Données
        gdf_temp = gdf.copy()
        gdf_temp['cluster'] = clusters_dict[n]
        gdf_temp = gdf_temp.sort_values(['cluster', 'nom_clean'])

        for idx, (_, row) in enumerate(gdf_temp.iterrows(), start=6):
            cluster = int(row['cluster'])

            # Commune
            ws.cell(row=idx, column=1, value=row['nom_clean']).border = border

            # Cluster (avec couleur cohérente)
            cell_cluster = ws.cell(row=idx, column=2, value=cluster)
            cell_cluster.border = border
            cell_cluster.alignment = Alignment(horizontal='center')
            cell_cluster.fill = PatternFill(
                start_color=CLUSTER_COLORS_EXCEL[cluster - 1],
                end_color=CLUSTER_COLORS_EXCEL[cluster - 1],
                fill_type="solid"
            )
            cell_cluster.font = Font(bold=True, color="000000")

            # Scores
            for col, val in enumerate([row['Score_Opp_0_10'], row['Score_Cho_0_10'],
                                       row['Score_Vec_0_10'], row['OppChoVec_0_10']], start=3):
                cell = ws.cell(row=idx, column=col, value=round(val, 2))
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

        # Ajuster largeurs
        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 12

        # Statistiques
        row_stats = len(gdf_temp) + 8
        ws.merge_cells(f'A{row_stats}:F{row_stats}')
        ws[f'A{row_stats}'] = "Statistiques par cluster"
        ws[f'A{row_stats}'].font = Font(bold=True, size=12)

        row_stats += 1
        for col, header in enumerate(['Cluster', 'N communes', 'Opp moyen', 'Cho moyen',
                                      'Vec moyen', 'OppChoVec moyen'], start=1):
            ws.cell(row=row_stats, column=col, value=header).font = Font(bold=True)

        for i in range(1, n + 1):
            row_stats += 1
            cluster_data = gdf_temp[gdf_temp['cluster'] == i]

            for col in range(1, 7):
                cell = ws.cell(row=row_stats, column=col)
                cell.fill = PatternFill(
                    start_color=CLUSTER_COLORS_EXCEL[i - 1],
                    end_color=CLUSTER_COLORS_EXCEL[i - 1],
                    fill_type="solid"
                )
                cell.font = Font(bold=True, color="000000")

            ws.cell(row=row_stats, column=1, value=f"Cluster {i}")
            ws.cell(row=row_stats, column=2, value=len(cluster_data))
            ws.cell(row=row_stats, column=3, value=round(cluster_data['Score_Opp_0_10'].mean(), 2))
            ws.cell(row=row_stats, column=4, value=round(cluster_data['Score_Cho_0_10'].mean(), 2))
            ws.cell(row=row_stats, column=5, value=round(cluster_data['Score_Vec_0_10'].mean(), 2))
            ws.cell(row=row_stats, column=6, value=round(cluster_data['OppChoVec_0_10'].mean(), 2))

        print(f"  [OK] Feuille '{n} clusters' creee ({len(gdf_temp)} communes)")

    wb.save(output_path)
    print(f"  [OK] Fichier Excel cree")

def exporter_vers_json(gdf, clusters_dict, output_dir):
    """Exporte vers JSON"""
    print(f"\n[INFO] Export vers JSON...")

    for n, clusters in clusters_dict.items():
        gdf_temp = gdf.copy()
        gdf_temp['cluster'] = clusters.astype(int)

        resultats = []
        for _, row in gdf_temp.iterrows():
            resultats.append({
                'commune': row['nom_clean'],
                'cluster': int(row['cluster']),
                'Score_Opp': float(row['Score_Opp_0_10']),
                'Score_Cho': float(row['Score_Cho_0_10']),
                'Score_Vec': float(row['Score_Vec_0_10']),
                'OppChoVec': float(row['OppChoVec_0_10'])
            })

        output_path = f"{output_dir}/cah_3d_{n}_clusters.json"
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(resultats, f, indent=2, ensure_ascii=False)

        print(f"  [OK] {output_path}")

def main():
    print("=" * 80)
    print("  REGENERATION COMPLETE CAH 3D - COULEURS COHERENTES")
    print("=" * 80)

    json_path = '../WEB/data_scores_0_10.json'
    geojson_path = '../WEB/Commune_Corse.geojson'
    output_dir = '../OUTPUT'
    dendrogramme_path = f'{output_dir}/cah_3d_dendrogramme.png'
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_excel = f'{output_dir}/cah_3d_resultats_{timestamp}.xlsx'
    output_communes = f'{output_dir}/clusters_3d_communes_{timestamp}.xlsx'

    data, gdf = charger_donnees(json_path, geojson_path)
    gdf_merged, X = preparer_donnees(data, gdf)
    X_scaled, Z = calculer_cah(X)
    generer_dendrogramme(Z, dendrogramme_path)
    clusters_dict = calculer_tous_clusters(Z, n_max=8)
    metriques_df = calculer_metriques_qualite(X_scaled, clusters_dict)
    generer_cartes_clusters(gdf_merged, clusters_dict, output_dir)
    analyses_dict = analyser_clusters(gdf_merged, clusters_dict)
    exporter_vers_excel(metriques_df, analyses_dict, output_excel)
    exporter_clusters_communes(gdf_merged, clusters_dict, output_communes)
    exporter_vers_json(gdf_merged, clusters_dict, output_dir)

    print("\n" + "=" * 80)
    print("RESUME DES METRIQUES DE QUALITE")
    print("=" * 80)
    print(metriques_df.to_string(index=False))

    best_silhouette = metriques_df.loc[metriques_df['silhouette'].idxmax()]
    best_davies = metriques_df.loc[metriques_df['davies_bouldin'].idxmin()]
    best_inertie = metriques_df.loc[metriques_df['inertie_intra'].idxmin()]

    print("\n" + "=" * 80)
    print("RECOMMANDATIONS")
    print("=" * 80)
    print(f"Meilleur Silhouette Score: {int(best_silhouette['n_clusters'])} clusters (score={best_silhouette['silhouette']:.4f})")
    print(f"Meilleur Davies-Bouldin: {int(best_davies['n_clusters'])} clusters (index={best_davies['davies_bouldin']:.4f})")
    print(f"Meilleure Inertie intra-classe: {int(best_inertie['n_clusters'])} clusters (inertie={best_inertie['inertie_intra']:.2f}, variance={best_inertie['variance_expliquee']:.1f}%)")

    print("\n" + "=" * 80)
    print("[OK] REGENERATION TERMINEE")
    print(f"     - Cartes: 7 fichiers PNG avec couleurs coherentes")
    print(f"     - Excel resultats: {output_excel}")
    print(f"     - Excel communes: {output_communes}")
    print(f"     - JSON: 7 fichiers")
    print("=" * 80)

if __name__ == "__main__":
    main()
