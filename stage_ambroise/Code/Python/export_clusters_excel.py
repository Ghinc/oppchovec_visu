"""
Génère un fichier Excel avec l'appartenance de chaque commune aux clusters
Une sheet par nombre de clusters (2 à 8)
Chaque sheet liste les communes avec leur cluster assigné
"""

import json
import pandas as pd
import geopandas as gpd
import numpy as np
from scipy.cluster.hierarchy import linkage, fcluster
from sklearn.preprocessing import StandardScaler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def charger_donnees(json_path, geojson_path):
    """Charge les données des indicateurs et le GeoJSON"""
    print("[INFO] Chargement des donnees...")

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    gdf = gpd.read_file(geojson_path)

    print(f"  [OK] {len(data)} communes chargees depuis JSON")
    print(f"  [OK] {len(gdf)} communes chargees depuis GeoJSON")

    return data, gdf

def preparer_donnees(data, gdf):
    """Prépare les données pour la CAH"""
    print("\n[INFO] Preparation des donnees pour CAH...")

    # Créer DataFrame avec les scores
    df_scores = pd.DataFrame.from_dict(data, orient='index')
    df_scores.index.name = 'commune'
    df_scores.reset_index(inplace=True)

    # Fusionner avec GeoDataFrame
    gdf['nom_clean'] = gdf['nom'].str.strip()
    gdf_merged = gdf.merge(df_scores, left_on='nom_clean', right_on='commune', how='inner')

    print(f"  [OK] {len(gdf_merged)} communes fusionnees avec succes")

    if len(gdf_merged) < len(data):
        print(f"  [WARN] {len(data) - len(gdf_merged)} communes n'ont pas pu etre fusionnees")
        # Identifier les communes manquantes
        communes_json = set(data.keys())
        communes_geo = set(gdf['nom_clean'])
        manquantes_geo = communes_json - communes_geo
        manquantes_json = communes_geo - communes_json
        if manquantes_geo:
            print(f"  [INFO] Communes dans JSON mais pas dans GeoJSON: {manquantes_geo}")
        if manquantes_json:
            print(f"  [INFO] Communes dans GeoJSON mais pas dans JSON: {manquantes_json}")

    # Extraire la variable OppChoVec normalisée 0-10
    X = gdf_merged[['OppChoVec_0_10']].values

    print(f"  [INFO] Matrice de donnees: {X.shape}")
    print(f"  [INFO] Min={X.min():.2f}, Max={X.max():.2f}, Moyenne={X.mean():.2f}")

    return gdf_merged, X

def calculer_clusters(X):
    """Calcule les clusters pour 2 à 8 groupes"""
    print("\n[INFO] Calcul de la CAH (methode Ward, distance euclidienne)...")

    # Standardiser
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    # CAH - Ward nécessite les observations, pas la matrice de distances
    Z = linkage(X_scaled, method='ward')

    # Calculer clusters pour 2 à 8
    clusters_dict = {}
    for n in range(2, 9):
        clusters = fcluster(Z, t=n, criterion='maxclust')
        clusters_dict[n] = clusters
        print(f"  [OK] {n} clusters calcules")

    return clusters_dict

def exporter_vers_excel(gdf, clusters_dict, output_path):
    """Exporte les appartenances aux clusters vers Excel"""
    print(f"\n[INFO] Creation du fichier Excel: {output_path}")

    wb = Workbook()
    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Palette de couleurs pour les clusters (8 couleurs distinctes)
    cluster_colors = [
        "E41A1C",  # Rouge
        "377EB8",  # Bleu
        "4DAF4A",  # Vert
        "984EA3",  # Violet
        "FF7F00",  # Orange
        "FFFF33",  # Jaune
        "A65628",  # Marron
        "F781BF"   # Rose
    ]

    # Créer une feuille par nombre de clusters
    for n in range(2, 9):
        ws = wb.create_sheet(title=f"{n} clusters")

        # Titre
        ws.merge_cells('A1:C1')
        ws['A1'] = f"Classification Hierarchique Ascendante - {n} clusters"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Info
        ws.merge_cells('A2:C2')
        ws['A2'] = f"Variable: OppChoVec (0-10) | Methode: Ward | Distance: Euclidienne"
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Date
        ws.merge_cells('A3:C3')
        ws['A3'] = f"Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M:%S')}"
        ws['A3'].font = Font(italic=True, size=9)
        ws['A3'].alignment = Alignment(horizontal='center')

        # En-têtes
        headers = ['Commune', 'Cluster', 'OppChoVec (0-10)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Préparer les données
        gdf_temp = gdf.copy()
        gdf_temp['cluster'] = clusters_dict[n]

        # Trier par cluster puis par nom de commune
        gdf_temp = gdf_temp.sort_values(['cluster', 'nom_clean'])

        # Écrire les données
        for idx, row in enumerate(gdf_temp.iterrows(), start=6):
            _, data = row
            commune = data['nom_clean']
            cluster = int(data['cluster'])
            oppchovec = data['OppChoVec_0_10']

            # Nom de commune
            cell_commune = ws.cell(row=idx, column=1, value=commune)
            cell_commune.border = border

            # Cluster (avec couleur de fond)
            cell_cluster = ws.cell(row=idx, column=2, value=cluster)
            cell_cluster.border = border
            cell_cluster.alignment = Alignment(horizontal='center')
            # Appliquer la couleur correspondant au cluster
            cell_cluster.fill = PatternFill(
                start_color=cluster_colors[cluster - 1],
                end_color=cluster_colors[cluster - 1],
                fill_type="solid"
            )
            cell_cluster.font = Font(bold=True, color="FFFFFF")

            # OppChoVec
            cell_opp = ws.cell(row=idx, column=3, value=round(oppchovec, 2))
            cell_opp.border = border
            cell_opp.alignment = Alignment(horizontal='center')

        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18

        # Ajouter statistiques par cluster
        row_stats = len(gdf_temp) + 8
        ws.merge_cells(f'A{row_stats}:C{row_stats}')
        ws[f'A{row_stats}'] = "Statistiques par cluster"
        ws[f'A{row_stats}'].font = Font(bold=True, size=12)

        row_stats += 1
        ws.cell(row=row_stats, column=1, value="Cluster").font = Font(bold=True)
        ws.cell(row=row_stats, column=2, value="N communes").font = Font(bold=True)
        ws.cell(row=row_stats, column=3, value="OppChoVec moyen").font = Font(bold=True)

        for i in range(1, n + 1):
            row_stats += 1
            cluster_data = gdf_temp[gdf_temp['cluster'] == i]
            ws.cell(row=row_stats, column=1, value=f"Cluster {i}")
            ws.cell(row=row_stats, column=2, value=len(cluster_data))
            ws.cell(row=row_stats, column=3, value=round(cluster_data['OppChoVec_0_10'].mean(), 2))

            # Appliquer la couleur du cluster
            for col in range(1, 4):
                cell = ws.cell(row=row_stats, column=col)
                cell.fill = PatternFill(
                    start_color=cluster_colors[i - 1],
                    end_color=cluster_colors[i - 1],
                    fill_type="solid"
                )
                cell.font = Font(bold=True, color="FFFFFF")

        print(f"  [OK] Feuille '{n} clusters' creee ({len(gdf_temp)} communes)")

    # Sauvegarder
    wb.save(output_path)
    print(f"\n  [OK] Fichier Excel cree avec succes")

def main():
    print("=" * 80)
    print("  EXPORT DES CLUSTERS - APPARTENANCE PAR COMMUNE")
    print("=" * 80)

    # Chemins
    json_path = '../WEB/data_scores_0_10.json'
    geojson_path = '../WEB/Commune_Corse.geojson'
    output_excel = '../OUTPUT/clusters_communes.xlsx'

    # Charger et préparer les données
    data, gdf = charger_donnees(json_path, geojson_path)
    gdf_merged, X = preparer_donnees(data, gdf)

    # Calculer les clusters
    clusters_dict = calculer_clusters(X)

    # Exporter vers Excel
    exporter_vers_excel(gdf_merged, clusters_dict, output_excel)

    print("\n" + "=" * 80)
    print(f"[OK] FICHIER EXCEL CREE: {output_excel}")
    print("     7 feuilles (2 a 8 clusters)")
    print(f"     {len(gdf_merged)} communes listees")
    print("=" * 80)

if __name__ == "__main__":
    main()
