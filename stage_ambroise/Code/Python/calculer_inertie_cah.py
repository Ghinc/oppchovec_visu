"""
Calcul de l'inertie intra-classe pour la Classification Hiérarchique Ascendante
Génère un graphique "elbow" pour aider à déterminer le nombre optimal de clusters

L'inertie intra-classe mesure la compacité des clusters :
- Plus l'inertie est faible, plus les clusters sont homogènes
- Le graphique "elbow" montre où l'ajout de clusters supplémentaires apporte peu d'amélioration
"""

import json
import pandas as pd
import geopandas as gpd
import numpy as np
import matplotlib.pyplot as plt
from scipy.cluster.hierarchy import linkage, fcluster
from sklearn.preprocessing import StandardScaler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def charger_donnees(json_path, geojson_path):
    """Charge les données des indicateurs et le GeoJSON"""
    print("[INFO] Chargement des donnees...")

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    gdf = gpd.read_file(geojson_path)

    print(f"  [OK] {len(data)} communes chargees depuis JSON")
    print(f"  [OK] {len(gdf)} communes chargees depuis GeoJSON")

    return data, gdf

def preparer_donnees(data, gdf):
    """Prépare les données pour la CAH"""
    print("\n[INFO] Preparation des donnees pour CAH...")

    # Créer DataFrame avec les scores
    df_scores = pd.DataFrame.from_dict(data, orient='index')
    df_scores.index.name = 'commune'
    df_scores.reset_index(inplace=True)

    # Fusionner avec GeoDataFrame
    gdf['nom_clean'] = gdf['nom'].str.strip()
    gdf_merged = gdf.merge(df_scores, left_on='nom_clean', right_on='commune', how='inner')

    print(f"  [OK] {len(gdf_merged)} communes fusionnees avec succes")

    # Extraire la variable OppChoVec normalisée 0-10
    X = gdf_merged[['OppChoVec_0_10']].values

    print(f"  [INFO] Matrice de donnees: {X.shape}")
    print(f"  [INFO] Min={X.min():.2f}, Max={X.max():.2f}, Moyenne={X.mean():.2f}")

    return gdf_merged, X

def calculer_clusters_et_inertie(X):
    """Calcule les clusters et l'inertie intra-classe pour 2 à 10 clusters"""
    print("\n[INFO] Calcul de la CAH et de l'inertie intra-classe...")

    # Standardiser
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    # CAH - Ward
    Z = linkage(X_scaled, method='ward')

    # Calculer pour 1 à 10 clusters
    n_clusters_range = range(1, 11)
    resultats = []

    for n in n_clusters_range:
        if n == 1:
            # Pour 1 cluster, tous les points sont dans le même cluster
            clusters = np.ones(len(X_scaled), dtype=int)
        else:
            clusters = fcluster(Z, t=n, criterion='maxclust')

        # Calculer l'inertie intra-classe
        inertie = 0.0
        for i in range(1, n + 1):
            # Points du cluster i
            cluster_points = X_scaled[clusters == i]
            if len(cluster_points) > 0:
                # Centroïde du cluster
                centroid = cluster_points.mean(axis=0)
                # Somme des distances au carré au centroïde
                inertie += np.sum((cluster_points - centroid) ** 2)

        # Calculer aussi l'inertie totale pour le pourcentage expliqué
        inertie_totale = np.sum((X_scaled - X_scaled.mean(axis=0)) ** 2)
        pct_inertie_expliquee = ((inertie_totale - inertie) / inertie_totale) * 100

        resultats.append({
            'n_clusters': n,
            'inertie_intra': inertie,
            'inertie_totale': inertie_totale,
            'pct_inertie_expliquee': pct_inertie_expliquee
        })

        print(f"  [OK] {n} cluster{'s' if n > 1 else ''}: "
              f"Inertie={inertie:.2f}, Variance expliquee={pct_inertie_expliquee:.2f}%")

    return pd.DataFrame(resultats)

def generer_graphique_elbow(df_inertie, output_path):
    """Génère le graphique elbow (coude) de l'inertie"""
    print(f"\n[INFO] Generation du graphique elbow: {output_path}")

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))

    # Graphique 1: Inertie intra-classe
    ax1.plot(df_inertie['n_clusters'], df_inertie['inertie_intra'],
             'o-', linewidth=2, markersize=8, color='#e74c3c')
    ax1.set_xlabel('Nombre de clusters', fontsize=12, weight='bold')
    ax1.set_ylabel('Inertie intra-classe', fontsize=12, weight='bold')
    ax1.set_title('Methode du coude (Elbow Method)\nInertie intra-classe',
                  fontsize=14, weight='bold')
    ax1.grid(True, alpha=0.3, linestyle='--')
    ax1.set_xticks(df_inertie['n_clusters'])

    # Ajouter les valeurs sur les points
    for idx, row in df_inertie.iterrows():
        ax1.annotate(f'{row["inertie_intra"]:.1f}',
                    xy=(row['n_clusters'], row['inertie_intra']),
                    xytext=(0, 10),
                    textcoords='offset points',
                    ha='center',
                    fontsize=9)

    # Graphique 2: Variance expliquée
    ax2.plot(df_inertie['n_clusters'], df_inertie['pct_inertie_expliquee'],
             'o-', linewidth=2, markersize=8, color='#3498db')
    ax2.set_xlabel('Nombre de clusters', fontsize=12, weight='bold')
    ax2.set_ylabel('Variance expliquee (%)', fontsize=12, weight='bold')
    ax2.set_title('Pourcentage de variance expliquee\npar la classification',
                  fontsize=14, weight='bold')
    ax2.grid(True, alpha=0.3, linestyle='--')
    ax2.set_xticks(df_inertie['n_clusters'])
    ax2.set_ylim(0, 100)

    # Ajouter les valeurs sur les points
    for idx, row in df_inertie.iterrows():
        ax2.annotate(f'{row["pct_inertie_expliquee"]:.1f}%',
                    xy=(row['n_clusters'], row['pct_inertie_expliquee']),
                    xytext=(0, 10),
                    textcoords='offset points',
                    ha='center',
                    fontsize=9)

    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  [OK] Graphique sauvegarde")

def calculer_diminution_inertie(df_inertie):
    """Calcule la diminution de l'inertie entre chaque étape"""
    print("\n[INFO] Calcul de la diminution de l'inertie...")

    df_inertie['diminution_inertie'] = df_inertie['inertie_intra'].diff().abs()
    df_inertie['pct_diminution'] = (df_inertie['diminution_inertie'] /
                                     df_inertie['inertie_intra'].shift(1) * 100)

    # Trouver le "coude" (où la diminution devient moins importante)
    # On cherche où le gain devient < 10% de l'inertie précédente
    coude = None
    for idx, row in df_inertie.iterrows():
        if idx > 0 and row['pct_diminution'] < 10:
            coude = row['n_clusters']
            break

    if coude:
        print(f"  [INFO] Coude detecte a {coude} clusters (diminution < 10%)")

    return df_inertie, coude

def exporter_vers_excel(df_inertie, output_path):
    """Exporte les résultats vers Excel"""
    print(f"\n[INFO] Creation du fichier Excel: {output_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Inertie intra-classe"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Titre
    ws.merge_cells('A1:F1')
    ws['A1'] = "Analyse de l'inertie intra-classe - Classification Hierarchique Ascendante"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Date
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M:%S')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(italic=True, size=10)

    # En-têtes
    headers = [
        'Nombre de clusters',
        'Inertie intra-classe',
        'Variance expliquee (%)',
        'Diminution inertie',
        'Diminution (%)'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Données
    for idx, row in df_inertie.iterrows():
        row_num = idx + 5
        ws.cell(row=row_num, column=1, value=row['n_clusters']).border = border
        ws.cell(row=row_num, column=2, value=round(row['inertie_intra'], 4)).border = border
        ws.cell(row=row_num, column=3, value=round(row['pct_inertie_expliquee'], 2)).border = border

        if pd.notna(row['diminution_inertie']):
            ws.cell(row=row_num, column=4, value=round(row['diminution_inertie'], 4)).border = border
        else:
            ws.cell(row=row_num, column=4, value='-').border = border

        if pd.notna(row['pct_diminution']):
            ws.cell(row=row_num, column=5, value=round(row['pct_diminution'], 2)).border = border
        else:
            ws.cell(row=row_num, column=5, value='-').border = border

    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18

    # Légende
    row_legende = len(df_inertie) + 7
    ws.merge_cells(f'A{row_legende}:E{row_legende}')
    ws[f'A{row_legende}'] = "Interpretation:"
    ws[f'A{row_legende}'].font = Font(bold=True, size=11)

    row_legende += 1
    ws[f'A{row_legende}'] = "• Inertie intra-classe:"
    ws[f'B{row_legende}'] = "Somme des distances au carre au sein des clusters (plus faible = meilleur)"

    row_legende += 1
    ws[f'A{row_legende}'] = "• Variance expliquee:"
    ws[f'B{row_legende}'] = "Pourcentage de variance capturee par la classification"

    row_legende += 1
    ws[f'A{row_legende}'] = "• Methode du coude:"
    ws[f'B{row_legende}'] = "Chercher le point ou l'ajout de clusters apporte peu d'amelioration"

    # Sauvegarder
    wb.save(output_path)
    print(f"  [OK] Fichier Excel cree")

def main():
    print("=" * 80)
    print("  CALCUL DE L'INERTIE INTRA-CLASSE - CAH")
    print("=" * 80)

    # Chemins
    json_path = '../WEB/data_scores_0_10.json'
    geojson_path = '../WEB/Commune_Corse.geojson'
    output_graph = '../OUTPUT/cah_inertie_elbow.png'
    output_excel = '../OUTPUT/cah_inertie.xlsx'

    # Charger et préparer les données
    data, gdf = charger_donnees(json_path, geojson_path)
    gdf_merged, X = preparer_donnees(data, gdf)

    # Calculer l'inertie
    df_inertie = calculer_clusters_et_inertie(X)

    # Calculer la diminution
    df_inertie, coude = calculer_diminution_inertie(df_inertie)

    # Générer le graphique
    generer_graphique_elbow(df_inertie, output_graph)

    # Exporter vers Excel
    exporter_vers_excel(df_inertie, output_excel)

    print("\n" + "=" * 80)
    print("RESUME DES RESULTATS")
    print("=" * 80)
    print("\nInertie intra-classe par nombre de clusters:")
    print(df_inertie[['n_clusters', 'inertie_intra', 'pct_inertie_expliquee', 'pct_diminution']].to_string(index=False))

    if coude:
        print(f"\n[INFO] Nombre optimal suggere par la methode du coude: {coude} clusters")

    print("\n" + "=" * 80)
    print(f"[OK] Graphique genere: {output_graph}")
    print(f"[OK] Fichier Excel cree: {output_excel}")
    print("=" * 80)

if __name__ == "__main__":
    main()
