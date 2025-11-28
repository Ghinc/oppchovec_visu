"""
Génère un fichier Excel avec l'appartenance de chaque commune aux clusters (CAH 3D)
Une sheet par nombre de clusters (2 à 8)
"""

import json
import pandas as pd
import geopandas as gpd
import numpy as np
from scipy.cluster.hierarchy import linkage, fcluster
from sklearn.preprocessing import StandardScaler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def charger_donnees(json_path, geojson_path):
    """Charge les données"""
    print("[INFO] Chargement des donnees...")

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    gdf = gpd.read_file(geojson_path)

    print(f"  [OK] {len(data)} communes chargees depuis JSON")
    print(f"  [OK] {len(gdf)} communes chargees depuis GeoJSON")

    return data, gdf

def preparer_donnees(data, gdf):
    """Prépare les données"""
    print("\n[INFO] Preparation des donnees pour CAH 3D...")

    df_scores = pd.DataFrame.from_dict(data, orient='index')
    df_scores.index.name = 'commune'
    df_scores.reset_index(inplace=True)

    gdf['nom_clean'] = gdf['nom'].str.strip()
    gdf_merged = gdf.merge(df_scores, left_on='nom_clean', right_on='commune', how='inner')

    print(f"  [OK] {len(gdf_merged)} communes fusionnees avec succes")

    X = gdf_merged[['Score_Opp_0_10', 'Score_Cho_0_10', 'Score_Vec_0_10']].values

    print(f"  [INFO] Matrice de donnees: {X.shape}")

    return gdf_merged, X

def calculer_clusters(X):
    """Calcule les clusters pour 2 à 8 groupes"""
    print("\n[INFO] Calcul de la CAH 3D (methode Ward, distance euclidienne)...")

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    Z = linkage(X_scaled, method='ward')

    clusters_dict = {}
    for n in range(2, 9):
        clusters = fcluster(Z, t=n, criterion='maxclust')
        clusters_dict[n] = clusters
        print(f"  [OK] {n} clusters calcules")

    return clusters_dict

def exporter_vers_excel(gdf, clusters_dict, output_path):
    """Exporte vers Excel"""
    print(f"\n[INFO] Creation du fichier Excel: {output_path}")

    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    cluster_colors = [
        "E41A1C", "377EB8", "4DAF4A", "984EA3",
        "FF7F00", "FFFF33", "A65628", "F781BF"
    ]

    for n in range(2, 9):
        ws = wb.create_sheet(title=f"{n} clusters")

        # Titre
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Classification Hierarchique Ascendante 3D - {n} clusters"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Info
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Variables: Opp, Cho, Vec | Methode: Ward | Distance: Euclidienne"
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Date
        ws.merge_cells('A3:F3')
        ws['A3'] = f"Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M:%S')}"
        ws['A3'].font = Font(italic=True, size=9)
        ws['A3'].alignment = Alignment(horizontal='center')

        # En-têtes
        headers = ['Commune', 'Cluster', 'Opp', 'Cho', 'Vec', 'OppChoVec']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Données
        gdf_temp = gdf.copy()
        gdf_temp['cluster'] = clusters_dict[n]
        gdf_temp = gdf_temp.sort_values(['cluster', 'nom_clean'])

        for idx, (_, row) in enumerate(gdf_temp.iterrows(), start=6):
            commune = row['nom_clean']
            cluster = int(row['cluster'])
            opp = row['Score_Opp_0_10']
            cho = row['Score_Cho_0_10']
            vec = row['Score_Vec_0_10']
            oppchovec = row['OppChoVec_0_10']

            # Commune
            cell_commune = ws.cell(row=idx, column=1, value=commune)
            cell_commune.border = border

            # Cluster (avec couleur)
            cell_cluster = ws.cell(row=idx, column=2, value=cluster)
            cell_cluster.border = border
            cell_cluster.alignment = Alignment(horizontal='center')
            cell_cluster.fill = PatternFill(
                start_color=cluster_colors[cluster - 1],
                end_color=cluster_colors[cluster - 1],
                fill_type="solid"
            )
            cell_cluster.font = Font(bold=True, color="FFFFFF")

            # Scores
            ws.cell(row=idx, column=3, value=round(opp, 2)).border = border
            ws.cell(row=idx, column=4, value=round(cho, 2)).border = border
            ws.cell(row=idx, column=5, value=round(vec, 2)).border = border
            ws.cell(row=idx, column=6, value=round(oppchovec, 2)).border = border

            ws.cell(row=idx, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=idx, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=idx, column=5).alignment = Alignment(horizontal='center')
            ws.cell(row=idx, column=6).alignment = Alignment(horizontal='center')

        # Ajuster largeurs
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15

        # Statistiques par cluster
        row_stats = len(gdf_temp) + 8
        ws.merge_cells(f'A{row_stats}:F{row_stats}')
        ws[f'A{row_stats}'] = "Statistiques par cluster"
        ws[f'A{row_stats}'].font = Font(bold=True, size=12)

        row_stats += 1
        headers_stats = ['Cluster', 'N communes', 'Opp moyen', 'Cho moyen', 'Vec moyen', 'OppChoVec moyen']
        for col, header in enumerate(headers_stats, start=1):
            ws.cell(row=row_stats, column=col, value=header).font = Font(bold=True)

        for i in range(1, n + 1):
            row_stats += 1
            cluster_data = gdf_temp[gdf_temp['cluster'] == i]
            ws.cell(row=row_stats, column=1, value=f"Cluster {i}")
            ws.cell(row=row_stats, column=2, value=len(cluster_data))
            ws.cell(row=row_stats, column=3, value=round(cluster_data['Score_Opp_0_10'].mean(), 2))
            ws.cell(row=row_stats, column=4, value=round(cluster_data['Score_Cho_0_10'].mean(), 2))
            ws.cell(row=row_stats, column=5, value=round(cluster_data['Score_Vec_0_10'].mean(), 2))
            ws.cell(row=row_stats, column=6, value=round(cluster_data['OppChoVec_0_10'].mean(), 2))

            for col in range(1, 7):
                cell = ws.cell(row=row_stats, column=col)
                cell.fill = PatternFill(
                    start_color=cluster_colors[i - 1],
                    end_color=cluster_colors[i - 1],
                    fill_type="solid"
                )
                cell.font = Font(bold=True, color="FFFFFF")

        print(f"  [OK] Feuille '{n} clusters' creee ({len(gdf_temp)} communes)")

    wb.save(output_path)
    print(f"\n  [OK] Fichier Excel cree avec succes")

def main():
    print("=" * 80)
    print("  EXPORT DES CLUSTERS 3D - APPARTENANCE PAR COMMUNE")
    print("=" * 80)

    json_path = '../WEB/data_scores_0_10.json'
    geojson_path = '../WEB/Commune_Corse.geojson'
    output_excel = '../OUTPUT/clusters_3d_communes.xlsx'

    data, gdf = charger_donnees(json_path, geojson_path)
    gdf_merged, X = preparer_donnees(data, gdf)
    clusters_dict = calculer_clusters(X)
    exporter_vers_excel(gdf_merged, clusters_dict, output_excel)

    print("\n" + "=" * 80)
    print(f"[OK] FICHIER EXCEL CREE: {output_excel}")
    print("     7 feuilles (2 a 8 clusters)")
    print(f"     {len(gdf_merged)} communes listees")
    print("=" * 80)

if __name__ == "__main__":
    main()
