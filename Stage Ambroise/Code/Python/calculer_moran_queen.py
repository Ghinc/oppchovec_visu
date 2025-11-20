"""
Calcul de l'indice de Moran pour les indicateurs OppChoVec, Opp, Cho et Vec
avec test de significativité à 1%

L'indice de Moran mesure l'autocorrélation spatiale des données.
- Valeur proche de 1 : forte autocorrélation positive (regroupement spatial)
- Valeur proche de 0 : absence d'autocorrélation (distribution aléatoire)
- Valeur proche de -1 : autocorrélation négative (dispersion)
"""

import json
import geopandas as gpd
import pandas as pd
from libpysal.weights import Queen, Rook
from esda.moran import Moran
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def charger_donnees(json_path, geojson_path):
    """Charge les données des indicateurs et le GeoJSON"""
    print("[INFO] Chargement des donnees...")

    # Charger les scores
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Charger le GeoJSON
    gdf = gpd.read_file(geojson_path)

    print(f"  [OK] {len(data)} communes chargees depuis JSON")
    print(f"  [OK] {len(gdf)} communes chargees depuis GeoJSON")

    return data, gdf

def preparer_geodataframe(data, gdf):
    """Prépare un GeoDataFrame avec les scores"""
    print("\n[INFO] Preparation du GeoDataFrame...")

    # Créer un DataFrame à partir des données JSON
    df_scores = pd.DataFrame.from_dict(data, orient='index')
    df_scores.index.name = 'commune'
    df_scores.reset_index(inplace=True)

    # Nettoyer les noms de communes dans le GeoJSON
    gdf['nom_clean'] = gdf['nom'].str.strip()

    # Fusionner avec le GeoDataFrame
    gdf_merged = gdf.merge(df_scores, left_on='nom_clean', right_on='commune', how='inner')

    print(f"  [OK] {len(gdf_merged)} communes fusionnees avec succes")

    if len(gdf_merged) < len(data):
        print(f"  [WARN] {len(data) - len(gdf_merged)} communes n'ont pas pu etre fusionnees")

    return gdf_merged

def creer_matrice_poids(gdf, type_contiguite='queen'):
    """Crée la matrice de poids spatiale"""
    print(f"\n[INFO] Creation de la matrice de poids spatiale ({type_contiguite})...")

    # FIX 1 : Reprojeter en Lambert-93 (EPSG:2154) pour distances correctes
    if gdf.crs and gdf.crs.to_string() != 'EPSG:2154':
        print(f"  [INFO] Reprojection de {gdf.crs.to_string()} vers Lambert-93 (EPSG:2154)...")
        gdf_proj = gdf.to_crs('EPSG:2154')
    else:
        gdf_proj = gdf

    if type_contiguite == 'queen':
        w = Queen.from_dataframe(gdf_proj, use_index=True)
    else:
        w = Rook.from_dataframe(gdf_proj, use_index=True)

    # FIX 2 : Row-standardiser la matrice de poids
    w.transform = 'R'

    print(f"  [OK] Matrice creee: {w.n} unites spatiales")
    print(f"  [INFO] Nombre moyen de voisins: {w.mean_neighbors:.2f}")
    print(f"  [INFO] Poids row-standardized")

    return w

def calculer_moran(gdf, w, variable_name):
    """Calcule l'indice de Moran pour une variable"""
    print(f"\n[INFO] Calcul de l'indice de Moran pour {variable_name}...")

    # FIX 3 : Gérer les NA - filtrer le GeoDataFrame et la matrice de poids
    gdf_clean = gdf[[variable_name, 'geometry']].dropna(subset=[variable_name])

    if len(gdf_clean) == 0:
        print(f"  [ERROR] Aucune valeur disponible pour {variable_name}")
        return None

    # Si des NA ont été retirés, il faut recréer la matrice de poids
    if len(gdf_clean) < len(gdf):
        print(f"  [INFO] {len(gdf) - len(gdf_clean)} valeurs NA retirees, recreation de la matrice...")
        gdf_clean_proj = gdf_clean.to_crs('EPSG:2154') if gdf_clean.crs.to_string() != 'EPSG:2154' else gdf_clean
        w_clean = Queen.from_dataframe(gdf_clean_proj, use_index=True)
        w_clean.transform = 'R'
        w = w_clean

    values = gdf_clean[variable_name].values
    moran = Moran(values, w)

    # Déterminer la significativité
    significatif_1pct = moran.p_sim < 0.01
    significatif_5pct = moran.p_sim < 0.05

    resultat = {
        'Variable': variable_name,
        'Indice_Moran_I': moran.I,
        'Esperance_E[I]': moran.EI,
        'Variance': moran.VI_norm,
        'Z_score': moran.z_norm,
        'P_value': moran.p_norm,
        'P_value_sim': moran.p_sim,
        'Significatif_1%': 'Oui' if significatif_1pct else 'Non',
        'Significatif_5%': 'Oui' if significatif_5pct else 'Non',
        'Interpretation': interpreter_moran(moran.I, significatif_1pct)
    }

    print(f"  [OK] I = {moran.I:.4f}")
    print(f"  [INFO] Z-score = {moran.z_norm:.4f}")
    print(f"  [INFO] P-value = {moran.p_norm:.6f}")
    print(f"  {'[OK] Significatif a 1%' if significatif_1pct else '[WARN] Non significatif a 1%'}")

    return resultat

def interpreter_moran(I, significatif):
    """Interprète l'indice de Moran"""
    if not significatif:
        return "Distribution aléatoire (non significatif)"
    elif I > 0.5:
        return "Forte autocorrélation positive (regroupement fort)"
    elif I > 0.3:
        return "Autocorrélation positive modérée (regroupement modéré)"
    elif I > 0.1:
        return "Faible autocorrélation positive (léger regroupement)"
    elif I > -0.1:
        return "Absence d'autocorrélation spatiale"
    elif I > -0.3:
        return "Faible autocorrélation négative (légère dispersion)"
    else:
        return "Autocorrélation négative (dispersion)"

def exporter_vers_excel(resultats, output_path):
    """Exporte les résultats vers Excel avec mise en forme"""
    print(f"\n[INFO] Creation du fichier Excel: {output_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Indices de Moran (Queen)"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    sig_fill_1pct = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    sig_fill_5pct = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Titre
    ws.merge_cells('A1:J1')
    ws['A1'] = "Indices de Moran - Autocorrelation Spatiale (Contiguite Queen)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Date
    ws.merge_cells('A2:J2')
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(italic=True, size=10)

    # En-têtes
    headers = [
        'Variable',
        'Indice de Moran (I)',
        'Espérance E[I]',
        'Z-score',
        'P-value (normale)',
        'P-value (simulation)',
        'Significatif à 1%',
        'Significatif à 5%',
        'Interprétation'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Données
    for row_idx, res in enumerate(resultats, start=5):
        ws.cell(row=row_idx, column=1, value=res['Variable']).border = border
        ws.cell(row=row_idx, column=2, value=round(res['Indice_Moran_I'], 4)).border = border
        ws.cell(row=row_idx, column=3, value=round(res['Esperance_E[I]'], 4)).border = border
        ws.cell(row=row_idx, column=4, value=round(res['Z_score'], 4)).border = border
        ws.cell(row=row_idx, column=5, value=round(res['P_value'], 6)).border = border
        ws.cell(row=row_idx, column=6, value=round(res['P_value_sim'], 6)).border = border

        # Cellule significativité 1%
        cell_1pct = ws.cell(row=row_idx, column=7, value=res['Significatif_1%'])
        cell_1pct.border = border
        if res['Significatif_1%'] == 'Oui':
            cell_1pct.fill = sig_fill_1pct
            cell_1pct.font = Font(bold=True)

        # Cellule significativité 5%
        cell_5pct = ws.cell(row=row_idx, column=8, value=res['Significatif_5%'])
        cell_5pct.border = border
        if res['Significatif_5%'] == 'Oui':
            cell_5pct.fill = sig_fill_5pct

        ws.cell(row=row_idx, column=9, value=res['Interpretation']).border = border

    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 50

    # Légende
    row_legende = len(resultats) + 7
    ws.merge_cells(f'A{row_legende}:I{row_legende}')
    ws[f'A{row_legende}'] = "Légende:"
    ws[f'A{row_legende}'].font = Font(bold=True, size=11)

    row_legende += 1
    ws[f'A{row_legende}'] = "• Indice de Moran (I):"
    ws[f'B{row_legende}'] = "Mesure l'autocorrélation spatiale (-1 à +1)"

    row_legende += 1
    ws[f'A{row_legende}'] = "• I > 0:"
    ws[f'B{row_legende}'] = "Autocorrélation positive (regroupement spatial)"

    row_legende += 1
    ws[f'A{row_legende}'] = "• I ≈ 0:"
    ws[f'B{row_legende}'] = "Distribution aléatoire (pas d'autocorrélation)"

    row_legende += 1
    ws[f'A{row_legende}'] = "• I < 0:"
    ws[f'B{row_legende}'] = "Autocorrélation négative (dispersion spatiale)"

    row_legende += 1
    ws[f'A{row_legende}'] = "• Z-score:"
    ws[f'B{row_legende}'] = "Standardisé (|Z| > 2.58 = significatif à 1%)"

    row_legende += 1
    ws[f'A{row_legende}'] = "• P-value:"
    ws[f'B{row_legende}'] = "Probabilité que le résultat soit dû au hasard"

    # Sauvegarder
    wb.save(output_path)
    print(f"  [OK] Fichier Excel cree avec succes")

def main():
    print("=" * 80)
    print("  CALCUL DES INDICES DE MORAN - CONTIGUITE QUEEN")
    print("=" * 80)

    # Chemins des fichiers
    json_path = '../WEB/data_scores_0_10.json'
    geojson_path = '../WEB/Commune_Corse.geojson'

    # Générer un nom unique pour éviter les conflits si le fichier est ouvert
    from datetime import datetime
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_excel = f'../OUTPUT/indices_moran_queen_{timestamp}.xlsx'

    # Charger les données
    data, gdf = charger_donnees(json_path, geojson_path)

    # Préparer le GeoDataFrame
    gdf_merged = preparer_geodataframe(data, gdf)

    # Créer la matrice de poids spatiale (Queen contiguity)
    w = creer_matrice_poids(gdf_merged, type_contiguite='queen')

    # Variables à analyser
    variables = ['OppChoVec', 'Score_Opp', 'Score_Cho', 'Score_Vec']

    # Calculer l'indice de Moran pour chaque variable
    resultats = []
    for var in variables:
        if var in gdf_merged.columns:
            resultat = calculer_moran(gdf_merged, w, var)
            if resultat:
                resultats.append(resultat)
        else:
            print(f"  [WARN] Variable {var} non trouvee dans les donnees")

    # Exporter vers Excel
    if resultats:
        exporter_vers_excel(resultats, output_excel)

        print("\n" + "=" * 80)
        print("RESUME DES RESULTATS")
        print("=" * 80)

        for res in resultats:
            print(f"\n{res['Variable']}:")
            print(f"  - Indice de Moran: {res['Indice_Moran_I']:.4f}")
            print(f"  - P-value: {res['P_value']:.6f}")
            print(f"  - Significatif a 1%: {res['Significatif_1%']}")
            print(f"  - Interpretation: {res['Interpretation']}")

        print("\n" + "=" * 80)
        print(f"[OK] Analyse terminee ! Fichier Excel: {output_excel}")
        print("=" * 80)
    else:
        print("\n[ERROR] Aucun resultat a exporter")

if __name__ == "__main__":
    main()
