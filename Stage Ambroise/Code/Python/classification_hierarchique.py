"""
lassification Hiérarchique Ascendante (CAH) sur OppChoVec
Teste de 2 à 8 clusters avec méthode de Ward et distance euclidienne

Génère :
- Dendrogramme
- Cartes des clusters par nombre de clusters
- Métriques de qualité (silhouette, Davies-Bouldin)
- Export JSON et Excel
"""

import json
import pandas as pd
import geopandas as gpd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy.cluster.hierarchy import linkage, dendrogram, fcluster
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import silhouette_score, davies_bouldin_score
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def charger_donnees(json_path, geojson_path):
    """Charge les données des indicateurs et le GeoJSON"""
    print("[INFO] Chargement des donnees...")

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    gdf = gpd.read_file(geojson_path)

    print(f"  [OK] {len(data)} communes chargees depuis JSON")
    print(f"  [OK] {len(gdf)} communes chargees depuis GeoJSON")

    return data, gdf

def preparer_donnees(data, gdf):
    """Prépare les données pour la CAH"""
    print("\n[INFO] Preparation des donnees pour CAH...")

    # Créer DataFrame avec les scores
    df_scores = pd.DataFrame.from_dict(data, orient='index')
    df_scores.index.name = 'commune'
    df_scores.reset_index(inplace=True)

    # Fusionner avec GeoDataFrame
    gdf['nom_clean'] = gdf['nom'].str.strip()
    gdf_merged = gdf.merge(df_scores, left_on='nom_clean', right_on='commune', how='inner')

    print(f"  [OK] {len(gdf_merged)} communes fusionnees avec succes")

    # Extraire la variable OppChoVec normalisée 0-10
    X = gdf_merged[['OppChoVec_0_10']].values

    print(f"  [INFO] Matrice de donnees: {X.shape}")
    print(f"  [INFO] Min={X.min():.2f}, Max={X.max():.2f}, Moyenne={X.mean():.2f}")

    return gdf_merged, X

def standardiser_donnees(X):
    """Standardise les données (moyenne=0, écart-type=1)"""
    print("\n[INFO] Standardisation des donnees...")

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    print(f"  [OK] Donnees standardisees")
    print(f"  [INFO] Nouvelle moyenne={X_scaled.mean():.6f}, ecart-type={X_scaled.std():.6f}")

    return X_scaled, scaler

def calculer_cah(X_scaled):
    """Calcule la Classification Hiérarchique Ascendante"""
    print("\n[INFO] Calcul de la CAH (methode Ward, distance euclidienne)...")

    # Calculer la liaison (linkage) directement sur les observations
    # Ward nécessite les observations, pas la matrice de distances
    Z = linkage(X_scaled, method='ward')
    print(f"  [OK] Matrice de liaison calculee")

    return Z

def generer_dendrogramme(Z, output_path):
    """Génère et sauvegarde le dendrogramme"""
    print(f"\n[INFO] Generation du dendrogramme: {output_path}")

    plt.figure(figsize=(16, 8))
    plt.title('Dendrogramme - Classification Hierarchique Ascendante (Ward)',
              fontsize=16, fontweight='bold')
    plt.xlabel('Communes', fontsize=12)
    plt.ylabel('Distance (critere de Ward)', fontsize=12)

    dendrogram(
        Z,
        leaf_rotation=90,
        leaf_font_size=6,
        color_threshold=None
    )

    plt.axhline(y=Z[-7, 2], color='r', linestyle='--', linewidth=1, label='8 clusters')
    plt.axhline(y=Z[-6, 2], color='orange', linestyle='--', linewidth=1, label='7 clusters')
    plt.axhline(y=Z[-5, 2], color='yellow', linestyle='--', linewidth=1, label='6 clusters')
    plt.axhline(y=Z[-4, 2], color='green', linestyle='--', linewidth=1, label='5 clusters')
    plt.axhline(y=Z[-3, 2], color='cyan', linestyle='--', linewidth=1, label='4 clusters')
    plt.axhline(y=Z[-2, 2], color='blue', linestyle='--', linewidth=1, label='3 clusters')
    plt.axhline(y=Z[-1, 2], color='purple', linestyle='--', linewidth=1, label='2 clusters')

    plt.legend(loc='upper right', fontsize=10)
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  [OK] Dendrogramme sauvegarde")

def calculer_clusters_multiples(Z, n_clusters_range):
    """Calcule les clusters pour plusieurs nombres"""
    print(f"\n[INFO] Calcul des clusters pour n={n_clusters_range}...")

    resultats = {}
    for n in n_clusters_range:
        clusters = fcluster(Z, t=n, criterion='maxclust')
        resultats[n] = clusters
        print(f"  [OK] {n} clusters: {len(np.unique(clusters))} groupes trouves")

    return resultats

def calculer_metriques_qualite(X_scaled, clusters_dict):
    """Calcule les métriques de qualité pour chaque nombre de clusters"""
    print("\n[INFO] Calcul des metriques de qualite...")

    metriques = []
    for n, clusters in clusters_dict.items():
        # Silhouette score (plus élevé = meilleur, range [-1, 1])
        silhouette = silhouette_score(X_scaled, clusters)

        # Davies-Bouldin index (plus faible = meilleur)
        davies_bouldin = davies_bouldin_score(X_scaled, clusters)

        # Inertie intra-classe (plus faible = meilleur)
        inertie_intra = 0.0
        for i in range(1, n + 1):
            cluster_points = X_scaled[clusters == i]
            if len(cluster_points) > 0:
                centroid = cluster_points.mean(axis=0)
                inertie_intra += np.sum((cluster_points - centroid) ** 2)

        # Variance expliquée
        inertie_totale = np.sum((X_scaled - X_scaled.mean(axis=0)) ** 2)
        pct_variance_expliquee = ((inertie_totale - inertie_intra) / inertie_totale) * 100

        metriques.append({
            'n_clusters': n,
            'silhouette': silhouette,
            'davies_bouldin': davies_bouldin,
            'inertie_intra': inertie_intra,
            'variance_expliquee': pct_variance_expliquee
        })

        print(f"  [OK] {n} clusters: Silhouette={silhouette:.4f}, Davies-Bouldin={davies_bouldin:.4f}, "
              f"Inertie={inertie_intra:.2f}, Variance={pct_variance_expliquee:.1f}%")

    return pd.DataFrame(metriques)

def generer_cartes_clusters(gdf, clusters_dict, output_dir):
    """Génère les cartes des clusters pour chaque nombre"""
    print("\n[INFO] Generation des cartes de clusters...")

    # Palette de couleurs distinctes
    colors_palette = ['#e41a1c', '#377eb8', '#4daf4a', '#984ea3',
                     '#ff7f00', '#ffff33', '#a65628', '#f781bf']

    for n, clusters in clusters_dict.items():
        gdf_plot = gdf.copy()
        gdf_plot['cluster'] = clusters

        fig, ax = plt.subplots(figsize=(12, 10))

        # Tracer chaque cluster avec une couleur différente
        legend_handles = []
        for i in range(1, n + 1):
            gdf_cluster = gdf_plot[gdf_plot['cluster'] == i]
            gdf_cluster.plot(ax=ax, color=colors_palette[i-1],
                           edgecolor='white', linewidth=0.5)
            # Créer un patch pour la légende
            legend_handles.append(mpatches.Patch(color=colors_palette[i-1],
                                                label=f'Cluster {i} (n={len(gdf_cluster)})'))

        ax.set_title(f'Classification Hierarchique Ascendante - {n} clusters\n' +
                    f'Variable: OppChoVec | Methode: Ward',
                    fontsize=14, fontweight='bold')
        ax.axis('off')

        # Légende
        ax.legend(handles=legend_handles, loc='upper left', fontsize=10,
                 frameon=True, fancybox=True, shadow=True)

        output_path = f"{output_dir}/cah_carte_{n}_clusters.png"
        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()

        print(f"  [OK] Carte sauvegardee: cah_carte_{n}_clusters.png")

def analyser_clusters(gdf, clusters_dict):
    """Analyse statistique des clusters"""
    print("\n[INFO] Analyse statistique des clusters...")

    analyses = {}
    for n, clusters in clusters_dict.items():
        gdf_temp = gdf.copy()
        gdf_temp['cluster'] = clusters

        stats = []
        for i in range(1, n + 1):
            cluster_data = gdf_temp[gdf_temp['cluster'] == i]['OppChoVec_0_10']
            stats.append({
                'cluster': i,
                'n_communes': len(cluster_data),
                'moyenne': cluster_data.mean(),
                'mediane': cluster_data.median(),
                'ecart_type': cluster_data.std(),
                'min': cluster_data.min(),
                'max': cluster_data.max()
            })

        analyses[n] = pd.DataFrame(stats)

    return analyses

def exporter_vers_excel(metriques_df, analyses_dict, output_path):
    """Exporte tous les résultats vers Excel"""
    print(f"\n[INFO] Export vers Excel: {output_path}")

    wb = Workbook()

    # Feuille 1 : Métriques de qualité
    ws1 = wb.active
    ws1.title = "Metriques de qualite"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    ws1['A1'] = "Metriques de qualite - Classification Hierarchique Ascendante"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:F1')

    headers = ['Nombre de clusters', 'Silhouette Score', 'Davies-Bouldin Index',
               'Inertie intra-classe', 'Variance expliquee (%)']
    for col, header in enumerate(headers, start=1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for idx, row in metriques_df.iterrows():
        ws1.cell(row=idx+4, column=1, value=row['n_clusters']).border = border
        ws1.cell(row=idx+4, column=2, value=round(row['silhouette'], 4)).border = border
        ws1.cell(row=idx+4, column=3, value=round(row['davies_bouldin'], 4)).border = border
        ws1.cell(row=idx+4, column=4, value=round(row['inertie_intra'], 2)).border = border
        ws1.cell(row=idx+4, column=5, value=round(row['variance_expliquee'], 2)).border = border

    # Trouver le meilleur selon silhouette
    best_silhouette = metriques_df.loc[metriques_df['silhouette'].idxmax()]
    ws1[f'A{len(metriques_df)+6}'] = f"Meilleur Silhouette Score: {best_silhouette['n_clusters']} clusters"
    ws1[f'A{len(metriques_df)+6}'].font = Font(bold=True, color="00B050")

    # Feuilles suivantes : Statistiques par nombre de clusters
    for n, stats_df in analyses_dict.items():
        ws = wb.create_sheet(title=f"{n} clusters")

        ws['A1'] = f"Statistiques descriptives - {n} clusters"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')

        headers = ['Cluster', 'N communes', 'Moyenne', 'Mediane', 'Ecart-type', 'Min', 'Max']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        for idx, row in stats_df.iterrows():
            ws.cell(row=idx+4, column=1, value=row['cluster']).border = border
            ws.cell(row=idx+4, column=2, value=row['n_communes']).border = border
            ws.cell(row=idx+4, column=3, value=round(row['moyenne'], 2)).border = border
            ws.cell(row=idx+4, column=4, value=round(row['mediane'], 2)).border = border
            ws.cell(row=idx+4, column=5, value=round(row['ecart_type'], 2)).border = border
            ws.cell(row=idx+4, column=6, value=round(row['min'], 2)).border = border
            ws.cell(row=idx+4, column=7, value=round(row['max'], 2)).border = border

    wb.save(output_path)
    print(f"  [OK] Fichier Excel cree")

def exporter_vers_json(gdf, clusters_dict, output_dir):
    """Exporte les résultats vers JSON pour le web"""
    print(f"\n[INFO] Export vers JSON...")

    for n, clusters in clusters_dict.items():
        gdf_temp = gdf.copy()
        gdf_temp['cluster'] = clusters

        data_dict = {}
        for idx, row in gdf_temp.iterrows():
            commune = row['nom_clean']
            data_dict[commune] = {
                'cluster': int(row['cluster']),
                'OppChoVec': float(row['OppChoVec_0_10'])
            }

        output_path = f"{output_dir}/cah_oppchovec_{n}_clusters.json"
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data_dict, f, indent=2, ensure_ascii=False)

        print(f"  [OK] {output_path}")

def main():
    print("=" * 80)
    print("  CLASSIFICATION HIERARCHIQUE ASCENDANTE - OPPCHOVEC")
    print("  Methode: Ward | Distance: Euclidienne | Clusters: 2-8")
    print("=" * 80)

    # Chemins
    json_path = '../WEB/data_scores_0_10.json'
    geojson_path = '../WEB/Commune_Corse.geojson'
    output_dir = '../OUTPUT'

    # Charger et préparer les données
    data, gdf = charger_donnees(json_path, geojson_path)
    gdf_merged, X = preparer_donnees(data, gdf)

    # Standardiser
    X_scaled, scaler = standardiser_donnees(X)

    # Calculer CAH
    Z = calculer_cah(X_scaled)

    # Générer dendrogramme
    generer_dendrogramme(Z, f"{output_dir}/cah_dendrogramme.png")

    # Calculer clusters pour 2 à 8
    n_clusters_range = range(2, 9)
    clusters_dict = calculer_clusters_multiples(Z, n_clusters_range)

    # Métriques de qualité
    metriques_df = calculer_metriques_qualite(X_scaled, clusters_dict)

    # Générer cartes
    generer_cartes_clusters(gdf_merged, clusters_dict, output_dir)

    # Analyser clusters
    analyses_dict = analyser_clusters(gdf_merged, clusters_dict)

    # Exporter Excel
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    exporter_vers_excel(metriques_df, analyses_dict,
                       f"{output_dir}/cah_resultats_{timestamp}.xlsx")

    # Exporter JSON
    exporter_vers_json(gdf_merged, clusters_dict, output_dir)

    print("\n" + "=" * 80)
    print("RESUME DES METRIQUES DE QUALITE")
    print("=" * 80)
    print(metriques_df.to_string(index=False))

    best_silhouette = metriques_df.loc[metriques_df['silhouette'].idxmax()]
    best_davies = metriques_df.loc[metriques_df['davies_bouldin'].idxmin()]
    best_inertie = metriques_df.loc[metriques_df['inertie_intra'].idxmin()]

    print("\n" + "=" * 80)
    print("RECOMMANDATIONS")
    print("=" * 80)
    print(f"Meilleur Silhouette Score: {int(best_silhouette['n_clusters'])} clusters (score={best_silhouette['silhouette']:.4f})")
    print(f"Meilleur Davies-Bouldin: {int(best_davies['n_clusters'])} clusters (index={best_davies['davies_bouldin']:.4f})")
    print(f"Meilleure Inertie intra-classe: {int(best_inertie['n_clusters'])} clusters (inertie={best_inertie['inertie_intra']:.2f}, variance={best_inertie['variance_expliquee']:.1f}%)")

    print("\n" + "=" * 80)
    print("[OK] TRAITEMENT TERMINE")
    print("=" * 80)

if __name__ == "__main__":
    main()
